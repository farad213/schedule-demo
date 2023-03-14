from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.http import FileResponse, HttpResponse, JsonResponse
from django.urls import reverse

from .models import Date, Project, DateBoundWithProject, Subproject, Artifact, Profile, SIT_with_date, SIT_project
from .forms import DateBoundWithProjectForm, ExportDates, SIT_with_date_form
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, wait
from io import BytesIO


def Monitoring_group_check(user):
    return user.groups.filter(name='Schedule - Monitoring').exists() or user.is_superuser or user.is_staff


@user_passes_test(Monitoring_group_check)
@login_required
def admin(request, year=datetime.date.year, month=datetime.date.month):
    if request.method == "GET":
        if "previous_month" in request.GET:
            if month > 1:
                month -= 1
            else:
                month = 12
                year -= 1

            return redirect(admin, year=year, month=month)
        elif "next_month" in request.GET:
            if month < 12:
                month += 1
            else:
                month = 1
                year += 1
            return redirect(admin, year=year, month=month)

    cal = get_calendar(year, month)
    last_month = datetime.date(year=year, month=month, day=15) - datetime.timedelta(30)
    next_month = datetime.date(year=year, month=month, day=15) + datetime.timedelta(30)
    dates_in_database = Date.objects.filter(date__range=(last_month, next_month))

    free_employees = dict()
    # appending Date model object to list
    for week in cal:
        for day in week:
            if Date.objects.filter(date=day[8]):
                date = Date.objects.get(date=day[8])
                day.append(date)

                all_possible_employees = User.objects.filter(groups__name="Schedule - Monitoring")
                employees_on_leave = date.employees_on_leave.all()
                available_employees = [employee.id for employee in all_possible_employees if
                                       employee not in employees_on_leave]

                projects = DateBoundWithProject.objects.filter(date=date)
                employees_working = set()
                for project in projects:
                    for employee in project.employee.all():
                        employees_working.add(employee.id)

                free_employees_on_date = [
                    f"{User.objects.get(pk=employee).first_name} {User.objects.get(pk=employee).last_name}" for employee
                    in available_employees if
                    employee not in employees_working]
                free_employees[day[3]] = free_employees_on_date

    for date in dates_in_database:

        if DateBoundWithProject.objects.filter(date=date):
            if date.state == "untouched":
                date.state = "done"
                date.save()
        else:
            date.state = "untouched"
            date.save()

        # if Date has DateBoundWithProject, updating Date with DateBoundWithProject's info
        if DateBoundWithProject.objects.filter(date=date):
            for project in DateBoundWithProject.objects.filter(date=date):
                for week in cal:
                    for day in week:
                        if day[8] == date.date:
                            if len(day) < 11:
                                day.append({})

                            if project.project.hasSubproject():
                                if project.project.project in day[10].keys():
                                    day[10][project.project.project].append(
                                        f"{project.subproject.name} - {len(project.profile.all())} cső")
                                else:
                                    day[10].update({project.project.project: [
                                        f"{project.subproject.name} - {len(project.profile.all())} cső"]})

                            else:
                                if project.project.project in day[10].keys():
                                    day[10][project.project.project].append(project.comment)
                                else:
                                    day[10].update({project.project.project: [project.comment]})

                            break
                        continue
                    continue

    dates_in_database = [date.__str__() for date in dates_in_database]
    month_str = str(month)

    today = datetime.datetime.today()
    start = today - datetime.timedelta(days=today.weekday())
    end = start + datetime.timedelta(days=6)
    export_dates = ExportDates(initial={"start": start, "end": end})
    context = {"cal": cal, "dates_in_database": dates_in_database, "year": year, "month_str": month_str,
               "export_dates": export_dates, "free_employees": free_employees, "employees": all_possible_employees}
    return render(request=request, template_name="schedule/admin.html", context=context)


def check_SIT(request):
    start = datetime.datetime.strptime(request.GET["start"], "%Y-%m-%d")
    end = datetime.datetime.strptime(request.GET["end"], "%Y-%m-%d")
    integritasvizsgalat_project = Project.objects.get(project="Integritásvizsgálat")
    integritasvizsgalat = DateBoundWithProject.objects.filter(project=integritasvizsgalat_project,
                                                              date__date__range=[start, end]).order_by("date__date")

    dates_without_sit_details = set()
    for sit in integritasvizsgalat:
        if not sit.sit_with_date_set.exists():
            dates_without_sit_details.add(sit.date.date.strftime("%Y/%m/%d"))
    dates_without_sit_details = sorted(dates_without_sit_details)

    return JsonResponse(dates_without_sit_details, safe=False)

@user_passes_test(Monitoring_group_check)
@login_required
def export_dates(request):
    start = datetime.datetime.strptime(request.GET["start"], "%Y-%m-%d")
    end = datetime.datetime.strptime(request.GET["end"], "%Y-%m-%d")

    sullyedesmeres_project = Project.objects.get(project="Süllyedésmérés")
    sullyedesmeres = DateBoundWithProject.objects.filter(project=sullyedesmeres_project,
                                                         date__date__range=[start, end]).order_by("date__date")
    wb = Workbook()

    # Süllyedésmérés
    ws = wb.active
    ws.title = "Süllyedésmérés"
    ws.append(["Iktatószám", "Megrendelő", "Helyszín", "Híd jele", "Szelvényszám", "Cső hossza", "Mérés dátuma"])

    breaklines = []
    row_no = 1
    for index, project in enumerate(sullyedesmeres):
        meres_datuma = project.date.date
        iktatoszam = project.subproject.subproject
        megrendelo = project.subproject.customer
        helyszin = project.subproject.name

        if index > 0 and meres_datuma != ws[f"G{row_no}"] and iktatoszam != ws[f"A{row_no}"]:
            breaklines.append(row_no)

        artifacts = None
        profiles = project.profile.all().order_by("profile")
        artifacts = {profile.artifact.artifact: [] for profile in profiles}
        sorted_artifact_keys = sorted(artifacts)

        for profile in profiles:
            artifacts[profile.artifact.artifact].append(profile)

        for artifact in sorted_artifact_keys:
            profiles = artifacts[artifact]
            hid_jele = artifact
            for profile in profiles:
                szelvenyszam = profile.profile
                cso_hossza = f"{profile.length} m"

                row_no += 1
                ws.append([iktatoszam, megrendelo, helyszin, hid_jele, szelvenyszam, cso_hossza, meres_datuma])

    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))

    thick_side = Side(style='thick', color="000000")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        new_style = ws[f"{col}1"].border.copy()
        new_style.top = thick_side
        new_style.bottom = thick_side
        ws[f"{col}1"].border = new_style

        new_style = ws[f"{col}{ws.max_row}"].border.copy()
        new_style.bottom = thick_side
        ws[f"{col}{ws.max_row}"].border = new_style

        for line in breaklines:
            new_style = ws[f"{col}{line}"].border.copy()
            new_style.bottom = thick_side
            ws[f"{col}{line}"].border = new_style

    for row in range(1, ws.max_row + 1):
        new_style = ws[f"A{row}"].border.copy()
        new_style.left = thick_side
        ws[f"A{row}"].border = new_style

        new_style = ws[f"G{row}"].border.copy()
        new_style.right = thick_side
        ws[f"G{row}"].border = new_style

    bold_font = Font(bold=True, name="Arial")
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = grey_fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 1.4)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    # Integritásvizsgálat
    ws = wb.create_sheet(title="Integritásvizsgálat")

    SITs = SIT_with_date.objects.filter(date_bound_with_project_object__date__date__range=[start, end]).order_by(
        "date_bound_with_project_object__date__date")

    ws.append(["Iktatószám", "Megrendelő", "Szerződés", "Helyszín", "Híd jele", "Támasz/Épület", "Cölöpök száma",
               "Mérés dátuma"])

    for sit in SITs:
        iktatoszam = sit.project_no.project_no
        megrendelo = sit.project_no.customer
        szerzodes = sit.project_no.contract
        helyszin = sit.project_no.location
        hid_jele = sit.bridge
        tamasz_epulet = sit.building
        no_of_piles = sit.no_of_piles
        meres_datuma = sit.date_bound_with_project_object.date.date

        ws.append(
            [iktatoszam, megrendelo, szerzodes, helyszin, hid_jele, tamasz_epulet, f"{no_of_piles} db", meres_datuma])

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = grey_fill

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = (max_length + 1.8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            new_style = ws[f"{col}1"].border.copy()
            new_style.top = thick_side
            new_style.bottom = thick_side
            ws[f"{col}1"].border = new_style

            new_style = ws[f"{col}{ws.max_row}"].border.copy()
            new_style.bottom = thick_side
            ws[f"{col}{ws.max_row}"].border = new_style

        for row in range(1, ws.max_row + 1):
            new_style = ws[f"A{row}"].border.copy()
            new_style.left = thick_side
            ws[f"A{row}"].border = new_style

            new_style = ws[f"H{row}"].border.copy()
            new_style.right = thick_side
            ws[f"H{row}"].border = new_style

    file_name = f"fugro_export_from_{request.GET['start']}_to_{request.GET['end']}.xlsx"
    buffer = BytesIO()
    executor = ThreadPoolExecutor()
    save_future = executor.submit(wb.save, buffer)
    wait([save_future])
    buffer.seek(0)

    response = FileResponse(buffer, as_attachment=True, filename=file_name)
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response


@user_passes_test(Monitoring_group_check)
@login_required
def date(request, year, month, day):
    date = Date.objects.get(date=datetime.date(year, month, day))
    projects = list(Project.objects.all())
    saved_projects_for_the_day = DateBoundWithProject.objects.filter(date=date).order_by("project")
    saved_projects = [saved_project for saved_project in saved_projects_for_the_day]
    untouched_projects = [project for project in projects]

    date_bound_project_form = DateBoundWithProjectForm(date=datetime.date(year, month, day))

    saved_project_forms = [DateBoundWithProjectForm(instance=project, date=datetime.date(year, month, day)) for project
                           in saved_projects_for_the_day]

    saved_projects_and_forms = zip(saved_projects, saved_project_forms)

    SIT_form = SIT_with_date_form()

    saved_sit_objects = list(SIT_with_date.objects.filter(
        date_bound_with_project_object__date__date=datetime.date(year, month, day)))

    saved_sit_forms = {object.date_bound_with_project_object.pk: SIT_with_date_form(instance=object) for object in
                       saved_sit_objects}

    if request.method == "POST":

        if "delete" in request.POST:
            # delete a project on that day
            project = Project.objects.filter(project=request.POST.get("delete")).get()
            project_to_delete = DateBoundWithProject.objects.filter(pk=int(request.POST.get("project_id")))
            project_to_delete.delete()
            return redirect("date", year, month, day)

        else:
            project = Project.objects.filter(project=request.POST.get("project_name")).get()
            filled_form = DateBoundWithProjectForm(data=request.POST, date=datetime.date(year, month, day))

            # create a project for that day
            if "untouched" in request.POST:
                # create a project for that day
                filled_model = filled_form.save(commit=False)
                filled_model.date = date
                filled_model.project = project
                filled_model.save()
                filled_form.save_m2m()
            # update a project for that day
            elif "saved" in request.POST:
                for saved_project in saved_projects_for_the_day:
                    if saved_project.id == int(request.POST.get("project_id")):
                        saved_project.employee.set(request.POST.getlist("employee"))
                        saved_project.vehicle.set(request.POST.getlist("vehicle"))
                        saved_project.comment = request.POST["comment"]
                        if "subproject" in request.POST or "artifact" in request.POST or "profile" in request.POST:
                            saved_project.subproject_id = request.POST["subproject"]
                            saved_project.artifact_id = request.POST["artifact"]
                            saved_project.profile.set(request.POST.getlist("profile"))
                        saved_project.save()

            return redirect("date", year, month, day)


    else:
        # changing dates within date.html being handled through GET requests
        if "next_day" in request.GET:
            that_day = datetime.date(year, month, day)
            next_day = that_day + datetime.timedelta(1)

            return redirect("date", next_day.year, next_day.month, next_day.day)
        elif "previous_day" in request.GET:
            that_day = datetime.date(year, month, day)
            previous_day = that_day + datetime.timedelta(-1)

            return redirect("date", previous_day.year, previous_day.month, previous_day.day)

    context = {"year_": year, "month_": month, "day_": day, "saved_projects_for_the_day": saved_projects_for_the_day,
               "untouched_projects": untouched_projects, "date_bound_project_form": date_bound_project_form,
               "saved_projects_and_forms": saved_projects_and_forms, "SIT_form": SIT_form,
               "saved_SIT_forms": saved_sit_forms}
    return render(request=request, template_name="schedule/date.html", context=context)


def artifacts(request):
    subproject_id = request.GET.get("subproject")
    artifacts = Artifact.objects.filter(subproject_id=subproject_id)
    context = {"artifacts": artifacts}
    return render(request, "schedule/ajax/artifacts.html", context)


def profiles(request):
    artifact_id = request.GET.get('artifact')
    profiles = Profile.objects.filter(artifact_id=artifact_id)
    context = {'profiles': profiles}
    return render(request, 'schedule/ajax/profiles.html', context)


def partial_save(request):
    project_id = int(request.GET.get("project_id"))
    project = DateBoundWithProject.objects.get(pk=project_id)

    # click on update button(can save profiles, employees, vehicles, comments)
    if not request.GET.get("r_type"):
        project.profile.add(*request.GET.getlist("profile[]"))
        if request.GET.get("employee[]"):
            project.employee.set(request.GET.getlist("employee[]"))
            project.save()
        if request.GET.get("vehicle[]"):
            project.vehicle.set(request.GET.getlist("vehicle[]"))
            project.save()
        if request.GET.get("comment"):
            project.comment = request.GET.get("comment")
            project.save()

    # clicks on add/remove icons
    elif request.GET.get("r_type") == "add or remove":
        id = request.GET.get("id")
        id_list = id.split("_")
        if id_list[0] == "add":
            project.profile.add(id_list[-1])
        elif id_list[0] == "remove":
            project.profile.remove(id_list[-1])

    # building context dict
    artifacts = set()
    subprojects = set()
    for profile in project.profile.all():
        artifacts.add(profile.artifact)
        subprojects.add(profile.artifact.subproject)
    context = {}
    for subproject in subprojects:
        context[subproject] = {}
    for artifact in artifacts:
        context[artifact.subproject].update({artifact: {}})
    for profile in project.profile.all().order_by("profile"):
        context[profile.artifact.subproject][profile.artifact].update({profile: "active"})
    for artifact in artifacts:
        for profile in artifact.profile_set.all().order_by("profile"):
            if profile not in context[artifact.subproject][artifact]:
                context[artifact.subproject][artifact].update({profile: "inactive"})
    return render(request, 'schedule/ajax/partial_save.html', context={"context": context})


def re_date_project(request):
    project = DateBoundWithProject.objects.get(pk=request.GET["project_id"])
    new_date = datetime.datetime.strptime(request.GET["new_date"], "%Y-%m-%d")
    project.date = Date.objects.get(date=new_date)
    project.save()

    year = request.GET["year"]
    month = request.GET["month"]
    day = request.GET["day"]

    return redirect(date, year, month, day)


@user_passes_test(Monitoring_group_check)
@login_required()
def user_calendar(request, year=datetime.date.year, month=datetime.date.month):
    if request.method == "GET":
        if "previous_month" in request.GET:
            if month > 1:
                month -= 1
            else:
                month = 12
                year -= 1

            return redirect(user_calendar, year=year, month=month)
        elif "next_month" in request.GET:
            if month < 12:
                month += 1
            else:
                month = 1
                year += 1
            return redirect(user_calendar, year=year, month=month)

    cal = get_calendar(year, month)

    user = request.user
    active_days = []
    for week in cal:
        for day in week:
            if Date.objects.filter(date=day[8]):
                date = Date.objects.get(date=day[8])
                projects_for_specific_day = DateBoundWithProject.objects.filter(date=date)
                for project in projects_for_specific_day:
                    if user in project.employee.all():
                        active_days.append(day[3])

    month_str = str(month)
    context = {"cal": cal, "year": year, "month_str": month_str, "active_days": active_days}
    return render(request=request, template_name="schedule/user_calendar.html", context=context)


@user_passes_test(Monitoring_group_check)
@login_required
def user_date(request, year, month, day):
    context_list = []
    user = request.user
    date = Date.objects.get(date=datetime.date(year, month, day))
    projects_for_that_day = DateBoundWithProject.objects.filter(date=date)
    for project in projects_for_that_day:
        if user in project.employee.all():
            project_name = project.project
            vehicles = project.vehicle.all()
            employees = project.employee.all()
            comment = project.comment
            if project.project.hasSubproject():
                artifacts = set()
                subprojects = set()
                for profile in project.profile.all():
                    artifacts.add(profile.artifact)
                    subprojects.add(profile.artifact.subproject)
                context = {}
                for subproject in subprojects:
                    context[subproject] = {}
                for artifact in artifacts:
                    context[artifact.subproject].update({artifact: {}})
                for profile in project.profile.all().order_by("profile"):
                    context[profile.artifact.subproject][profile.artifact].update({profile: "active"})
                for artifact in artifacts:
                    for profile in artifact.profile_set.all().order_by("profile"):
                        if profile not in context[artifact.subproject][artifact]:
                            context[artifact.subproject][artifact].update({profile: "inactive"})
                mini_context = {"project_name": project_name, "vehicles": vehicles, "employees": employees,
                                "comment": comment, "hasSubproject": context}
            else:
                mini_context = {"project_name": project_name, "vehicles": vehicles, "employees": employees,
                                "comment": comment}
            context_list.append(mini_context)
    context = {"context": context_list, "date": date}

    return render(request=request, template_name="schedule/calendar_date.html", context=context)


@login_required
def vacation(request):
    dates_when_user_is_on_leave = Date.objects.filter(employees_on_leave=request.user)
    dates_when_user_is_on_leave = [date.date.strftime("%Y-%m-%d") for date in dates_when_user_is_on_leave]
    context = {"dates_when_user_is_on_leave": dates_when_user_is_on_leave}
    return render(request=request, template_name="schedule/vacation.html", context=context)


def vacation_set_ajax(request):
    dates = request.GET.getlist('dates[]')
    dates_when_user_is_on_leave = Date.objects.filter(employees_on_leave=request.user)
    date_objs_when_user_is_on_leave = [date.date for date in dates_when_user_is_on_leave]
    dates_from_calendar = []
    bad_dates_already_working = []
    for date in dates:
        date = date.split("-")
        date_obj = datetime.date(int(date[0]), int(date[1]), int(date[2]))
        if Date.objects.filter(date=date_obj):
            date_db = Date.objects.get(date=date_obj)
            if DateBoundWithProject.objects.filter(date=date_db, employee=request.user):
                bad_dates_already_working.append(date_obj)
                continue

        dates_from_calendar.append(date_obj)
        if date_obj in date_objs_when_user_is_on_leave:
            pass
        else:
            if Date.objects.filter(date=date_obj):
                date_db = Date.objects.get(date=date_obj)
                date_db.employees_on_leave.add(request.user)
    dates_to_delete = [date for date in date_objs_when_user_is_on_leave if date not in dates_from_calendar]
    for date in dates_to_delete:
        Date.objects.get(date=date).employees_on_leave.remove(request.user)

    dates_when_user_is_on_leave = Date.objects.filter(employees_on_leave=request.user)
    dates_when_user_is_on_leave = [date.date.strftime("%Y-%m-%d") for date in dates_when_user_is_on_leave]
    print(bad_dates_already_working, "bad")
    print(dates_when_user_is_on_leave, "good")
    context = {"dates_when_user_is_on_leave": dates_when_user_is_on_leave,
               "bad_dates_already_working": bad_dates_already_working}
    return render(request=request, template_name="schedule/ajax/vacation_ajax.html", context=context)


def get_calendar(year, month):
    assert isinstance(year, int)
    assert isinstance(month, int)

    import calendar

    day_dict = {0: "Hétfő", 1: "Kedd", 2: "Szerda", 3: "Csütörtök", 4: "Péntek", 5: "Szombat", 6: "Vasárnap"}

    c = calendar.Calendar()
    dates = c.monthdatescalendar(year, month)

    for j, week in enumerate(dates):
        for i, date in enumerate(week):
            full_date_str = date.strftime("%Y.%m.%d")
            date_str = date.strftime("%m/%d")
            day_int = date.day
            day = day_dict[i]
            month_str = date.strftime("%m")
            month_int = date.month
            year_int = date.year
            month_str_from_int = str(date.month)
            week[i] = [date_str, day, month_str, full_date_str, day_int, month_int, year_int, month_str_from_int, date]
    return dates

@user_passes_test(Monitoring_group_check)
@login_required
def repeat_project(request):
    year = int(request.GET["year"])
    month = int(request.GET["month"])
    day = int(request.GET["day"])
    project_id = request.GET["project_id"]
    repeat = request.GET["repeat"]
    date = datetime.date(year, month, day)
    project = DateBoundWithProject.objects.get(pk=int(project_id))

    dates = []
    if repeat == "Soha":
        pass
    elif repeat == "Egy hétig":
        for index in range(1, 7):
            date_to_check = date + datetime.timedelta(index)
            if date_to_check.weekday() in [0, 1, 2, 3, 4]:
                dates.append(date_to_check)

    elif repeat == "Hetente":
        for index in [7, 14, 21, 28]:
            date_to_check = date + datetime.timedelta(index)
            # if weekday
            if date_to_check.weekday() in [0, 1, 2, 3, 4]:
                dates.append(date_to_check)
            # if Sunday
            elif date + datetime.timedelta(index + 1).weekday() in [0, 1, 2, 3, 4]:
                dates.append(date + datetime.timedelta(index + 1))
            # if Saturday
            else:
                dates.append(date + datetime.timedelta(index + 2))

    elif repeat == "Kéthetente":
        for index in [14, 28, 42, 56]:
            date_to_check = date + datetime.timedelta(index)
            # if weekday
            if date_to_check.weekday() in [0, 1, 2, 3, 4]:
                dates.append(date_to_check)
            # if Sunday
            elif date + datetime.timedelta(index + 1).weekday() in [0, 1, 2, 3, 4]:
                dates.append(date + datetime.timedelta(index + 1))
            # if Saturday
            else:
                dates.append(date + datetime.timedelta(index + 2))

    elif repeat == "Havonta":
        # every 28 days
        for index in [28, 56]:
            date_to_check = date + datetime.timedelta(index)
            # if weekday
            if date_to_check.weekday() in [0, 1, 2, 3, 4]:
                dates.append(date_to_check)
            # if Sunday
            elif date + datetime.timedelta(index + 1).weekday() in [0, 1, 2, 3, 4]:
                dates.append(date + datetime.timedelta(index + 1))
            # if Saturday
            else:
                dates.append(date + datetime.timedelta(index + 2))

    for new_date in dates:
        if Date.objects.filter(date=date):
            if new_date.weekday() in [0, 1, 2, 3, 4]:
                new_date = Date.objects.get(date=new_date)
                new = DateBoundWithProject.objects.create(project=project.project,
                                                          date=new_date,
                                                          comment=project.comment)
                new.employee.set(project.employee.all())
                new.vehicle.set(project.vehicle.all())
                if project.subproject:
                    new.subproject = project.subproject
                    new.artifact = project.artifact
                    new.profile.set(project.profile.all())
                new.save()

    return redirect("date", year, month, day)

@user_passes_test(Monitoring_group_check)
@login_required
def SIT_details(request):
    year = int(request.GET["year"])
    month = int(request.GET["month"])
    day = int(request.GET["day"])
    project_id = request.GET["project_id"]
    project = DateBoundWithProject.objects.get(pk=int(project_id))
    if SIT_with_date.objects.filter(date_bound_with_project_object=project):
        sit_object = SIT_with_date.objects.get(date_bound_with_project_object=project)
        sit_object.bridge = request.GET["bridge"]
        sit_object.building = request.GET["building"]
        sit_object.no_of_piles = int(request.GET["no_of_piles"])
        sit_object.project_no = SIT_project.objects.get(pk=request.GET["project_no"])
        sit_object.save()
    else:
        form = SIT_with_date_form(request.GET)
        if form.is_valid():
            sit_object = form.save(commit=False)
            sit_object.date_bound_with_project_object = project
            sit_object.save()

    return redirect("date", year, month, day)


def user_selection(request, year=datetime.date.today().year, month=datetime.date.today().month, id=None):
    print(id)
    if "id" in request.GET.keys():
        if "Admin" in request.GET["id"]:
            return HttpResponse(reverse("admin", args=[year, month]))
        else:
            id = request.GET["id"]
    user = User.objects.get(pk=int(id))

    cal = get_calendar(year, month)

    active_days = []
    for week in cal:
        for day in week:
            if Date.objects.filter(date=day[8]):
                date = Date.objects.get(date=day[8])
                projects_for_specific_day = DateBoundWithProject.objects.filter(date=date)
                for project in projects_for_specific_day:
                    if user in project.employee.all():
                        active_days.append(day[3])

    month_str = str(month)
    context = {"cal": cal, "year": year, "month_str": month_str, "active_days": active_days, "id": id}
    return render(request=request, template_name="schedule/ajax/user_selection.html", context=context)


def user_selection_date(request, year, month, day, id):
    context_list = []
    user = User.objects.get(pk=id)
    date = Date.objects.get(date=datetime.date(year, month, day))
    projects_for_that_day = DateBoundWithProject.objects.filter(date=date)
    for project in projects_for_that_day:
        if user in project.employee.all():
            project_name = project.project
            vehicles = project.vehicle.all()
            employees = project.employee.all()
            comment = project.comment
            if project.project.hasSubproject():
                artifacts = set()
                subprojects = set()
                for profile in project.profile.all():
                    artifacts.add(profile.artifact)
                    subprojects.add(profile.artifact.subproject)
                context = {}
                for subproject in subprojects:
                    context[subproject] = {}
                for artifact in artifacts:
                    context[artifact.subproject].update({artifact: {}})
                for profile in project.profile.all().order_by("profile"):
                    context[profile.artifact.subproject][profile.artifact].update({profile: "active"})
                for artifact in artifacts:
                    for profile in artifact.profile_set.all().order_by("profile"):
                        if profile not in context[artifact.subproject][artifact]:
                            context[artifact.subproject][artifact].update({profile: "inactive"})
                mini_context = {"project_name": project_name, "vehicles": vehicles, "employees": employees,
                                "comment": comment, "hasSubproject": context}
            else:
                mini_context = {"project_name": project_name, "vehicles": vehicles, "employees": employees,
                                "comment": comment}
            context_list.append(mini_context)
    context = {"context": context_list, "date": date, "id": id}

    return render(request=request, template_name="schedule/calendar_date.html", context=context)


def month_change_user_selection(request):
    year = int(request.GET["year"])
    month = int(request.GET["month"])
    id = int(request.GET["id"])
    date = datetime.date(year, month, 1)
    if request.GET["direction"] == "forwards":
        new_date = date + datetime.timedelta(40)
    else:
        new_date = date - datetime.timedelta(1)
    year = new_date.year
    month = new_date.month

    return user_selection(request, year, month, id)


def day_nav_user_date(request):
    date = request.GET["date"]
    date = [int(item) for item in date.split(".")]
    date = datetime.date(date[0], date[1], date[2])

    if request.GET["day_selection"] == "forwards":
        date = date + datetime.timedelta(1)
    else:
        date = date - datetime.timedelta(1)

    if request.GET["user_id"]:
        id = request.GET["user_id"]

        return user_selection_date(request, date.year, date.month, date.day, id)

    return user_date(request, date.year, date.month, date.day)

